from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PACK = ROOT / "Atlas_Rover_Mk1_制造图纸包_V1.0"


REQUIRED_FILES = [
    "说明.md",
    "采购清单_Atlas_Rover_Mk1_V1.0.csv",
    "采购清单_Atlas_Rover_Mk1_V1.0.md",
    "采购执行表_Atlas_Rover_Mk1_V1.0.xlsx",
    "采买执行清单_Atlas_Rover_Mk1_V1.0.md",
    "规格核对_Atlas_Rover_Mk1_V1.0.md",
    "装配指南_Atlas_Rover_Mk1_V1.0.md",
    "接线说明_Atlas_Rover_Mk1_V1.0.md",
    "详细组装与接线手册_Atlas_Rover_Mk1_V1.0.md",
    "程序设计_Atlas_Rover_Mk1_V1.0.md",
    "语音与双板UART控制方案_Atlas_Rover_Mk1_V1.0.md",
    "组装模拟与效果确认_Atlas_Rover_Mk1_V1.0.md",
    "一致性复审报告_Atlas_Rover_Mk1_V1.0.md",
    "资料来源_Atlas_Rover_Mk1_V1.0.md",
    "图纸/AR-MK1-001_总体尺寸.svg",
    "图纸/AR-MK1-002_层级结构.svg",
    "图纸/AR-MK1-003_底板模板.svg",
    "图纸/AR-MK1-004_接线框图.svg",
    "输出/PDF/Atlas_Rover_Mk1_制造图纸包_V1.0.pdf",
]

REQUIRED_TERMS = [
    "DRV8833 双路 H 桥电机驱动板（必备）",
    "PCA9685 I2C PWM 扩展板（推荐）",
    "WS2812B 灯条或灯环（外接车灯模块）",
    "WS2812B 数据电平转换器（推荐）",
    "M2 铜柱套装",
    "黄铜丝/黄铜棒",
    "Seeed Studio XIAO ESP32C3 底盘控制板",
]

TEXT_FILES = [
    "说明.md",
    "采购清单_Atlas_Rover_Mk1_V1.0.md",
    "采买执行清单_Atlas_Rover_Mk1_V1.0.md",
    "规格核对_Atlas_Rover_Mk1_V1.0.md",
    "装配指南_Atlas_Rover_Mk1_V1.0.md",
    "接线说明_Atlas_Rover_Mk1_V1.0.md",
    "详细组装与接线手册_Atlas_Rover_Mk1_V1.0.md",
    "程序设计_Atlas_Rover_Mk1_V1.0.md",
    "语音与双板UART控制方案_Atlas_Rover_Mk1_V1.0.md",
    "组装模拟与效果确认_Atlas_Rover_Mk1_V1.0.md",
    "一致性复审报告_Atlas_Rover_Mk1_V1.0.md",
]


def fail(message: str) -> None:
    raise SystemExit(f"FAIL: {message}")


def read_text(rel: str) -> str:
    return (PACK / rel).read_text(encoding="utf-8")


def main() -> None:
    missing = [rel for rel in REQUIRED_FILES if not (PACK / rel).exists()]
    if missing:
        fail("missing files: " + ", ".join(missing))

    csv_path = PACK / "采购清单_Atlas_Rover_Mk1_V1.0.csv"
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    expected_columns = {"推荐搜索词", "京东", "淘宝", "拼多多"}
    missing_columns = expected_columns - set(rows[0].keys())
    if missing_columns:
        fail("BOM missing purchase link columns: " + ", ".join(sorted(missing_columns)))
    bom_names = {row["名称"] for row in rows}
    for term in REQUIRED_TERMS:
        if term not in bom_names:
            fail(f"BOM missing required term: {term}")
    for row in rows:
        if not row["推荐搜索词"]:
            fail(f"BOM missing search query for: {row['名称']}")
        for platform in ["京东", "淘宝", "拼多多"]:
            if not row[platform].startswith("https://"):
                fail(f"BOM missing {platform} link for: {row['名称']}")

    xlsx_path = PACK / "采购执行表_Atlas_Rover_Mk1_V1.0.xlsx"
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    ws = wb.active
    headers = [ws.cell(4, col).value for col in range(1, 9)]
    if headers != ["项目", "数量", "优先级", "推荐搜索词", "下单校验", "京东", "淘宝", "拼多多"]:
        fail("purchase execution workbook headers are not aligned")
    if ws.cell(5, 1).value != "ESP32-S3-DualEye-Touch-LCD-1.28":
        fail("purchase execution workbook missing core board row")
    if not str(ws.cell(5, 6).value).startswith("https://search.jd.com/"):
        fail("purchase execution workbook missing JD URL for core board")

    all_text = "\n".join(read_text(rel) for rel in TEXT_FILES)
    for term in REQUIRED_TERMS:
        if term not in all_text:
            fail(f"text docs missing required term: {term}")

    required_phrases = [
        "PCA9685 不是电机驱动",
        "WS2812B 数据电平",
        "AHCT/HCT",
        "不要把 5 V 升压输出接进主控板电池接口",
        "用 M2，不用 M3",
        "不承担主结构",
        "可以做出接近参考图的三层黄铜笼架效果",
        "Pin 9 为 UART_RXD",
        "Pin 10 为 UART_TXD",
        "DualEye LCD1 Pin10 UART_TXD  -> XIAO D7 / GPIO20 / RX",
        "XIAO D2 / GPIO4",
        "电机供电不要从 DualEye 板上取",
        "每条运动命令都必须带 `duration_ms`",
        "AR1,STOP",
        "不带 `AR1,` 前缀",
        "双目表情程序",
        "多主题页面",
        "MimiClaw",
        "RoverIntent",
        "不能直接写 UART",
    ]
    for phrase in required_phrases:
        if phrase not in all_text:
            fail(f"missing safety/consistency phrase: {phrase}")

    if "WS2812B 数据线接 PCA9685" not in all_text:
        fail("missing warning that WS2812B data should not connect to PCA9685")

    print("PASS: Atlas Rover Mk.1 manufacturing pack consistency checks passed")


if __name__ == "__main__":
    main()
